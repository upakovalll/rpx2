"""
Direct Excel writers using openpyxl for precise control over output.
This module provides writers that bypass pandas' limitations with duplicate column names.
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import pandas as pd
from typing import List, Dict, Any, Optional
from io import BytesIO
import logging

from .definitions.column_order import AUDIT_COLUMN_ORDER
from .formatters import ExcelFormatter

logger = logging.getLogger(__name__)


class DirectExcelWriter:
    """Write Excel files directly using openpyxl for full control over formatting."""
    
    @staticmethod
    def write_audit_format_sheet(
        workbook: Workbook, 
        sheet_name: str, 
        data: pd.DataFrame,
        apply_formatting: bool = True
    ) -> Worksheet:
        """
        Write a sheet in the 87-column audit format with exact column names.
        This method preserves duplicate column names exactly as specified.
        
        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name for the worksheet
            data: DataFrame with data (columns may have __posXX suffixes)
            apply_formatting: Whether to apply Excel formatting
            
        Returns:
            The created worksheet
        """
        # Create or get worksheet
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)
        
        # Write headers exactly as specified in AUDIT_COLUMN_ORDER
        # This preserves duplicate column names
        for col_idx, header in enumerate(AUDIT_COLUMN_ORDER, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            
            if apply_formatting:
                # Apply header formatting
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Check if data has position suffixes
        has_position_suffixes = any('__pos' in str(col) for col in data.columns)
        
        # Write data rows
        for row_idx, row_data in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                # Convert value to Excel-compatible format
                excel_value = ExcelFormatter.safe_value(value)
                ws.cell(row=row_idx, column=col_idx, value=excel_value)
        
        # Auto-adjust column widths if formatting is enabled
        if apply_formatting:
            DirectExcelWriter._adjust_column_widths(ws, data, AUDIT_COLUMN_ORDER)
        
        return ws
    
    @staticmethod
    def write_standard_sheet(
        workbook: Workbook,
        sheet_name: str,
        data: pd.DataFrame,
        apply_formatting: bool = True
    ) -> Worksheet:
        """
        Write a standard sheet (no duplicate columns).
        
        Args:
            workbook: openpyxl Workbook object
            sheet_name: Name for the worksheet
            data: DataFrame with data
            apply_formatting: Whether to apply Excel formatting
            
        Returns:
            The created worksheet
        """
        # Create or get worksheet
        if sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
        else:
            ws = workbook.create_sheet(sheet_name)
        
        # Write headers
        for col_idx, header in enumerate(data.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=str(header))
            
            if apply_formatting:
                # Apply header formatting
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Write data rows
        for row_idx, row_data in enumerate(data.itertuples(index=False), 2):
            for col_idx, value in enumerate(row_data, 1):
                # Convert value to Excel-compatible format
                excel_value = ExcelFormatter.safe_value(value)
                ws.cell(row=row_idx, column=col_idx, value=excel_value)
        
        # Auto-adjust column widths if formatting is enabled
        if apply_formatting:
            DirectExcelWriter._adjust_column_widths(ws, data, list(data.columns))
        
        return ws
    
    @staticmethod
    def _adjust_column_widths(ws: Worksheet, data: pd.DataFrame, headers: List[str]) -> None:
        """
        Adjust column widths based on content.
        
        Args:
            ws: Worksheet to adjust
            data: DataFrame with data
            headers: List of header names
        """
        for idx, header in enumerate(headers, 1):
            column_letter = get_column_letter(idx)
            max_length = len(str(header))
            
            # Check data values for max length (sample first 100 rows)
            for row in ws.iter_rows(min_row=2, max_row=min(101, ws.max_row), 
                                   min_col=idx, max_col=idx):
                for cell in row:
                    try:
                        value_length = len(str(cell.value or ''))
                        if value_length > max_length:
                            max_length = value_length
                    except:
                        pass
            
            # Set column width (max 50)
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def create_workbook_with_sheets(sheets_data: List[Dict[str, Any]]) -> BytesIO:
        """
        Create a complete Excel workbook with multiple sheets.
        
        Args:
            sheets_data: List of dictionaries with sheet information:
                - 'name': Sheet name
                - 'data': DataFrame with data
                - 'format': 'audit' or 'standard'
                - 'apply_formatting': Boolean (optional, default True)
                
        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Add each sheet
        for sheet_info in sheets_data:
            sheet_name = sheet_info['name']
            data = sheet_info['data']
            format_type = sheet_info.get('format', 'standard')
            apply_formatting = sheet_info.get('apply_formatting', True)
            
            if format_type == 'audit':
                DirectExcelWriter.write_audit_format_sheet(
                    wb, sheet_name, data, apply_formatting
                )
            else:
                DirectExcelWriter.write_standard_sheet(
                    wb, sheet_name, data, apply_formatting
                )
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output


class AuditFormatExcelWriter:
    """Specialized writer for 87-column audit format Excel files."""
    
    @staticmethod
    def create_excel_file(
        loans_data: Optional[List[Dict[str, Any]]] = None,
        pricing_data: Optional[List[Dict[str, Any]]] = None,
        properties_data: Optional[List[Dict[str, Any]]] = None,
        metadata: Optional[Dict[str, Any]] = None
    ) -> BytesIO:
        """
        Create an Excel file in the 87-column audit format.
        
        Args:
            loans_data: Loan records for the Loans sheet
            pricing_data: Pricing records for the Pricing Results sheet
            properties_data: Property records for the Properties sheet
            metadata: Metadata for the Report Info sheet
            
        Returns:
            BytesIO object containing the Excel file
        """
        from .sheets.pricing_sheet import PricingSheet
        
        sheets_to_create = []
        
        # Add metadata sheet if provided
        if metadata:
            from datetime import datetime
            metadata.setdefault('Generated', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            metadata.setdefault('Report', 'RPX Portfolio Analysis - 87 Column Audit Format')
            
            metadata_items = [{'Field': k, 'Value': v} for k, v in metadata.items()]
            metadata_df = pd.DataFrame(metadata_items)
            
            sheets_to_create.append({
                'name': 'Report Info',
                'data': metadata_df,
                'format': 'standard'
            })
        
        # Add loans sheet if provided
        if loans_data:
            # Use PricingSheet to format the data
            loans_df = PricingSheet._format_pricing_data(loans_data)
            sheets_to_create.append({
                'name': 'Loans',
                'data': loans_df,
                'format': 'audit'
            })
        
        # Add pricing results sheet if provided
        if pricing_data:
            # Use PricingSheet to format the data
            pricing_df = PricingSheet._format_pricing_data(pricing_data)
            sheets_to_create.append({
                'name': 'Pricing Results',
                'data': pricing_df,
                'format': 'audit'
            })
        
        # Add properties sheet if provided
        if properties_data:
            properties_df = pd.DataFrame(properties_data)
            # Apply safe value conversion
            for col in properties_df.columns:
                properties_df[col] = properties_df[col].apply(ExcelFormatter.safe_value)
            
            sheets_to_create.append({
                'name': 'Properties',
                'data': properties_df,
                'format': 'standard'
            })
        
        # Create the workbook
        return DirectExcelWriter.create_workbook_with_sheets(sheets_to_create)