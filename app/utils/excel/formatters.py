"""
Excel formatting and data conversion utilities.
Handles data type conversions and Excel-specific formatting.
"""

from typing import Any, Dict, List, Optional
from datetime import date, datetime
from decimal import Decimal
from uuid import UUID
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .column_definitions import (
    NUMERIC_COLUMNS,
    PERCENTAGE_COLUMNS,
    BASIS_POINTS_COLUMNS,
    DATE_COLUMNS
)


class ExcelFormatter:
    """Utilities for Excel data formatting and conversion."""
    
    @staticmethod
    def safe_value(value: Any) -> Any:
        """
        Convert database types to Excel-compatible values, preserving data types.
        
        Args:
            value: Raw value from database
            
        Returns:
            Excel-compatible value
        """
        # Check for pandas NA first
        if pd.isna(value):
            return None
        elif isinstance(value, UUID):
            return str(value)
        elif isinstance(value, Decimal):
            return float(value)
        elif isinstance(value, datetime):
            # Remove timezone info for Excel compatibility
            if value.tzinfo is not None:
                return value.replace(tzinfo=None)
            return value
        elif isinstance(value, date):
            return value
        elif value is None:
            return None  # Return None instead of empty string to preserve Excel data type
        return value
    
    @staticmethod
    def convert_dataframe_types(df: pd.DataFrame) -> pd.DataFrame:
        """
        Apply safe_value conversion to all columns in a DataFrame.
        
        Args:
            df: DataFrame to convert
            
        Returns:
            DataFrame with converted values
        """
        for col in df.columns:
            df[col] = df[col].apply(ExcelFormatter.safe_value)
        return df
    
    @staticmethod
    def format_percentage(value: Any, as_decimal: bool = False) -> Optional[float]:
        """
        Format percentage values.
        
        Args:
            value: Raw percentage value
            as_decimal: If True, return as decimal (0.05 for 5%)
            
        Returns:
            Formatted percentage
        """
        if value is None or pd.isna(value):
            return None
        
        try:
            numeric_value = float(value)
            if as_decimal:
                # When exporting as decimal, treat values >=1 as percentages
                return numeric_value / 100 if numeric_value >= 1 else numeric_value
            # Otherwise ensure decimal inputs (0-1) are converted to percentage scale
            return numeric_value * 100 if numeric_value <= 1 else numeric_value
        except (TypeError, ValueError):
            return None
    
    @staticmethod
    def format_basis_points(value: Any, to_decimal: bool = False) -> Optional[float]:
        """
        Format basis points values.
        
        Args:
            value: Raw basis points value
            to_decimal: If True, convert to decimal (250 bps -> 0.025)
            
        Returns:
            Formatted value
        """
        if value is None or pd.isna(value):
            return None
        
        try:
            numeric_value = float(value)
            if to_decimal:
                return numeric_value / 10000
            return numeric_value
        except (TypeError, ValueError):
            return None
    
    @staticmethod
    def format_currency(value: Any) -> Optional[float]:
        """
        Format currency values.
        
        Args:
            value: Raw currency value
            
        Returns:
            Formatted currency value
        """
        if value is None or pd.isna(value):
            return None
        
        try:
            return float(value)
        except (TypeError, ValueError):
            return None
    
    @staticmethod
    def format_worksheet(worksheet: Worksheet, dataframe: pd.DataFrame) -> None:
        """
        Apply Excel formatting to a worksheet.
        
        Args:
            worksheet: openpyxl worksheet object
            dataframe: DataFrame used to create the worksheet
        """
        # Header formatting
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Apply header formatting
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Auto-adjust column widths
        for idx, col in enumerate(dataframe.columns):
            column_letter = get_column_letter(idx + 1)
            max_length = len(str(col))
            
            # Check data values for max length
            for row_idx, value in enumerate(dataframe[col]):
                try:
                    value_length = len(str(value))
                    if value_length > max_length:
                        max_length = value_length
                except:
                    pass
            
            # Set column width (max 50)
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def apply_column_formatting(worksheet: Worksheet, column_index: int, column_name: str) -> None:
        """
        Apply specific formatting based on column type.
        
        Args:
            worksheet: openpyxl worksheet
            column_index: 1-based column index
            column_name: Name of the column for type identification
        """
        column_letter = get_column_letter(column_index)
        
        # Determine format based on column type
        if column_name in PERCENTAGE_COLUMNS:
            # Apply percentage format
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                cell.number_format = '0.00%'
        
        elif column_name in BASIS_POINTS_COLUMNS:
            # Apply number format for basis points
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                cell.number_format = '#,##0'
        
        elif column_name in DATE_COLUMNS:
            # Apply date format
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                cell.number_format = 'yyyy-mm-dd'
        
        elif column_name in NUMERIC_COLUMNS:
            # Apply number format with commas
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{column_letter}{row}"]
                if 'Balance' in column_name or 'Value' in column_name:
                    cell.number_format = '#,##0.00'
                else:
                    cell.number_format = '#,##0.0000'


class DataProcessor:
    """Process raw data for Excel export."""
    
    @staticmethod
    def prepare_loan_data(raw_data: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        Prepare loan data for Excel export.
        
        Args:
            raw_data: Raw loan data from database
            
        Returns:
            Processed DataFrame
        """
        if not raw_data:
            return pd.DataFrame()
        
        df = pd.DataFrame(raw_data)
        
        # Apply safe value conversion
        df = ExcelFormatter.convert_dataframe_types(df)
        
        # Apply specific formatting for known columns
        if 'rpx_base_spread_bps' in df.columns:
            df['rpx_base_spread_bps'] = df['rpx_base_spread_bps'].apply(
                lambda x: ExcelFormatter.format_basis_points(x, to_decimal=False)
            )
        
        if 'pd' in df.columns:
            df['pd'] = df['pd'].apply(
                lambda x: ExcelFormatter.format_percentage(x, as_decimal=True)
            )
        
        if 'lgd' in df.columns:
            df['lgd'] = df['lgd'].apply(
                lambda x: ExcelFormatter.format_percentage(x, as_decimal=True)
            )
        
        return df
    
    @staticmethod
    def prepare_market_data(benchmarks: List[Dict], spreads: List[Dict]) -> pd.DataFrame:
        """
        Prepare market data (benchmarks and spreads) for Excel export.
        
        Args:
            benchmarks: Benchmark rate data
            spreads: Credit spread data
            
        Returns:
            Combined market data DataFrame
        """
        market_data = []
        
        # Process benchmarks
        if benchmarks:
            for bench in benchmarks:
                market_data.append({
                    'Type': 'Benchmark',
                    'Name': bench.get('benchmark_type', ''),
                    'Tenor': bench.get('tenor', ''),
                    'Rate': ExcelFormatter.format_percentage(bench.get('rate', 0)),
                    'Effective Date': bench.get('effective_date', '')
                })
        
        # Process spreads
        if spreads:
            for spread in spreads:
                market_data.append({
                    'Type': 'Credit Spread',
                    'Name': f"{spread.get('property_sector', '')} - {spread.get('term_bucket', '')}",
                    'Tenor': '',
                    'Rate': ExcelFormatter.format_basis_points(
                        spread.get('spread_bps', 0), to_decimal=True
                    ),
                    'Effective Date': spread.get('effective_date', '')
                })
        
        if market_data:
            return pd.DataFrame(market_data)
        return pd.DataFrame()


class CellStyler:
    """Apply conditional formatting and cell styles."""
    
    @staticmethod
    def highlight_negative_values(worksheet: Worksheet, column_index: int, start_row: int = 2):
        """
        Highlight negative values in red.
        
        Args:
            worksheet: openpyxl worksheet
            column_index: 1-based column index
            start_row: Row to start checking (default 2, after header)
        """
        column_letter = get_column_letter(column_index)
        red_font = Font(color="FF0000")
        
        for row in range(start_row, worksheet.max_row + 1):
            cell = worksheet[f"{column_letter}{row}"]
            try:
                if cell.value and float(cell.value) < 0:
                    cell.font = red_font
            except (TypeError, ValueError):
                pass
    
    @staticmethod
    def apply_alternating_rows(worksheet: Worksheet, start_row: int = 2):
        """
        Apply alternating row colors for better readability.
        
        Args:
            worksheet: openpyxl worksheet
            start_row: Row to start alternating (default 2, after header)
        """
        light_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        for row in range(start_row, worksheet.max_row + 1):
            if (row - start_row) % 2 == 1:
                for cell in worksheet[row]:
                    cell.fill = light_fill
    
    @staticmethod
    def add_borders(worksheet: Worksheet):
        """
        Add borders to all cells with data.
        
        Args:
            worksheet: openpyxl worksheet
        """
        from openpyxl.styles import Border, Side
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, 
                                      min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.border = thin_border